from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from models import Partner, ParseResult
from parsers.base import BaseParser
from parsers.table_tools import build_column_map, find_header_index, rows_to_items
from utils.text import clean_text


class XlsxParser(BaseParser):
    file_format = "xlsx"

    def parse(self, path: Path, partner: Partner | None = None) -> ParseResult:
        partner = self.make_partner(path, partner)
        document = self.make_document(path, partner)
        result = ParseResult(document=document)
        workbook_path = self._ensure_xlsx(path, result)
        if workbook_path is None:
            return result.finalize()
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            result.items.extend(self._parse_sheet(sheet, document))
        result.document.raw_content = self._sample_workbook_text(workbook)
        if not result.items:
            result.errors.append("Workbook has no recognized price rows")
        return result.finalize()

    def _parse_sheet(self, sheet, document) -> list:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        header_index = find_header_index(rows)
        if header_index is None:
            return []
        columns = build_column_map(rows, header_index)
        return rows_to_items(rows, document, columns, header_index + 1, sheet.title)

    def _ensure_xlsx(self, path: Path, result: ParseResult) -> Path | None:
        if path.suffix.lower() == ".xlsx":
            return path
        # Старый бинарный .xls: сначала пробуем xlrd (чистый Python, в зависимостях),
        # затем — LibreOffice как тяжёлый фолбэк.
        converted = self._convert_xls_with_xlrd(path) or self._convert_xls_with_libreoffice(path)
        if converted:
            return converted
        result.errors.append("Cannot read .xls: install xlrd or LibreOffice")
        return None

    @staticmethod
    def _convert_xls_with_xlrd(path: Path) -> Path | None:
        """Прочитать .xls через xlrd и пересохранить в .xlsx для общего пути парсинга."""
        try:
            import xlrd
            from openpyxl import Workbook
        except ImportError:
            return None
        try:
            book = xlrd.open_workbook(str(path))
        except Exception:
            return None
        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        for sheet in book.sheets():
            ws = out_wb.create_sheet(title=(sheet.name or "Sheet")[:31])
            for r in range(sheet.nrows):
                ws.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        temp_dir = Path(tempfile.mkdtemp(prefix="medarchive_xls_"))
        output = temp_dir / f"{path.stem}.xlsx"
        out_wb.save(output)
        return output

    @staticmethod
    def _convert_xls_with_libreoffice(path: Path) -> Path | None:
        soffice = shutil.which("libreoffice") or shutil.which("soffice")
        if not soffice:
            return None
        temp_dir = Path(tempfile.mkdtemp(prefix="medarchive_xls_"))
        command = [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(temp_dir), str(path)]
        completed = subprocess.run(command, capture_output=True, text=True, timeout=60)
        if completed.returncode != 0:
            return None
        output = temp_dir / f"{path.stem}.xlsx"
        return output if output.exists() else None

    @staticmethod
    def _sample_workbook_text(workbook) -> str:
        chunks: list[str] = []
        for sheet in workbook.worksheets:
            chunks.append(f"# {sheet.title}")
            for row in sheet.iter_rows(max_row=30, values_only=True):
                chunks.append(" | ".join(clean_text(cell) for cell in row if clean_text(cell)))
        return "\n".join(chunks)
