"""Локальный тестер пайплайна на СВОИХ данных — без БД и без загрузчика.

Парсит твои прайсы и сразу сопоставляет позиции с твоим справочником,
печатает отчёт (auto / review / unmatched + % автосопоставления) и пишет CSV.

Примеры:
    # справочник из JSON, прайсы из папки/zip/одного файла
    python try_pipeline.py --archive ./my_prices.zip --services ./catalog.json

    # справочник из XLSX (колонки: service_name, synonyms, category)
    python try_pipeline.py --archive ./price.xlsx --services ./catalog.xlsx

    # справочник взять из живой базы Supabase (repository.load_services)
    python try_pipeline.py --archive ./price.xlsx --from-db

    # покрутить пороги
    python try_pipeline.py --archive ./price.xlsx --services ./catalog.json --auto 0.85 --unmatched 0.65

Формат JSON-справочника: список объектов
    [{"service_id": "...", "service_name": "...", "synonyms": ["..."], "category": "..."}]
"""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from archive_parser import ArchiveParser
from config import ParserConfig
from normalizer import Normalizer, ServiceRecord
from normalizer.config import NormalizerConfig


# ---- загрузка справочника из разных источников ---------------------------
def services_from_json(path: Path) -> list[ServiceRecord]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [_record(d) for d in data]


# ключевые слова для авто-поиска колонок (по подстроке, регистр не важен)
_NAME_KEYS = ("наимен", "назван", "услуг", "service", "name")
_SYN_KEYS = ("синоним", "synonym", "альтерн")
_CAT_KEYS = ("категор", "category", "раздел", "группа", "group", "тип")
_ID_KEYS = ("service_id", "код", "id", "артикул")


def services_from_xlsx(path: Path, verbose: bool = True) -> list[ServiceRecord]:
    """Грузит справочник из XLSX, сам находя строку заголовка по всем листам."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    best: list[ServiceRecord] = []
    best_info = ""
    for sheet in wb.worksheets:
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        header_idx, mapping = _find_header(rows)
        if header_idx is None:
            continue
        records = _rows_to_services(rows, header_idx, mapping)
        if len(records) > len(best):
            best = records
            cols = ", ".join(f"{k}={rows[header_idx][v]!r}" for k, v in mapping.items() if v is not None)
            best_info = f"лист '{sheet.title}', заголовок в строке {header_idx + 1}: {cols}"
    if verbose:
        if best:
            print(f"Справочник распознан: {best_info}")
            print("  примеры:", "; ".join(s.service_name for s in best[:3]))
        else:
            print("ВНИМАНИЕ: заголовок справочника не найден ни на одном листе.")
            _dump_first_rows(wb)
    return best


def _find_header(rows: list[list], max_scan: int = 30):
    """Ищет строку, где есть колонка с названием услуги. Возвращает (индекс, маппинг)."""
    for i, row in enumerate(rows[:max_scan]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        i_name = _match_col(cells, _NAME_KEYS)
        if i_name is None:
            continue
        return i, {
            "name": i_name,
            "synonyms": _match_col(cells, _SYN_KEYS),
            "category": _match_col(cells, _CAT_KEYS),
            "id": _match_col(cells, _ID_KEYS),
        }
    return None, {}


def _match_col(cells: list[str], keys: tuple[str, ...]):
    for idx, cell in enumerate(cells):
        if cell and any(k in cell for k in keys):
            return idx
    return None


def _rows_to_services(rows: list[list], header_idx: int, m: dict) -> list[ServiceRecord]:
    out: list[ServiceRecord] = []
    i_name, i_syn, i_cat, i_id = m["name"], m.get("synonyms"), m.get("category"), m.get("id")
    for r in rows[header_idx + 1:]:
        name = r[i_name] if i_name < len(r) else None
        if not name or not str(name).strip():
            continue
        syn = r[i_syn] if i_syn is not None and i_syn < len(r) else None
        out.append(ServiceRecord(
            service_id=str(r[i_id]) if i_id is not None and i_id < len(r) and r[i_id] else str(name),
            service_name=str(name).strip(),
            synonyms=_split_synonyms(syn),
            category=str(r[i_cat]).strip() if i_cat is not None and i_cat < len(r) and r[i_cat] else None,
        ))
    return out


def _dump_first_rows(wb, n: int = 8) -> None:
    """Печатает первые строки каждого листа — чтобы вручную увидеть структуру."""
    for sheet in wb.worksheets:
        print(f"  --- лист '{sheet.title}' (первые {n} строк) ---")
        for i, row in enumerate(sheet.iter_rows(max_row=n, values_only=True)):
            print(f"   [{i}]", [str(c)[:25] if c is not None else "" for c in row][:8])


def services_from_db() -> list[ServiceRecord]:
    import repository

    return [_record(row) for row in repository.load_services()]


def _record(d: dict) -> ServiceRecord:
    return ServiceRecord(
        service_id=str(d.get("service_id") or d["service_name"]),
        service_name=d["service_name"],
        synonyms=list(d.get("synonyms") or []),
        category=d.get("category"),
    )


def _split_synonyms(value) -> list[str]:
    if not value:
        return []
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("["):
            try:
                return list(json.loads(text))
            except json.JSONDecodeError:
                pass
        return [s.strip() for s in text.split(";") if s.strip()]
    return list(value)


# ---- основной прогон ------------------------------------------------------
def run(args) -> None:
    services = _load_services(args)
    print(f"Справочник: {len(services)} услуг")
    if args.inspect_services:
        return  # только проверка справочника, без парсинга архива

    payload = ArchiveParser(ParserConfig()).parse(args.archive)
    items = payload["items"]
    rep = payload["report"]
    print(f"Парсер: документов {rep['documents_total']} "
          f"(ошибок {rep['documents_error']}), позиций {rep['items_total']}")

    cfg = NormalizerConfig()
    if args.auto is not None:
        cfg.auto_threshold = args.auto
    if args.unmatched is not None:
        cfg.unmatched_threshold = args.unmatched
    if args.no_embeddings:
        cfg.use_embeddings = False
    cfg.validate()
    print(f"Эмбеддинги: {'ВКЛ (LaBSE)' if cfg.use_embeddings else 'ВЫКЛ (быстрый прогон)'}")
    norm = Normalizer(services, cfg)

    rows, counts = _match_all(norm, items)
    _print_report(counts, len(items), cfg)
    _write_csv(Path(args.out), rows)
    print(f"\nДетализация по позициям: {args.out}")


def _load_services(args) -> list[ServiceRecord]:
    if args.from_db:
        return services_from_db()
    if not args.services:
        raise SystemExit("Укажи --services путь (JSON/XLSX) или --from-db")
    path = Path(args.services)
    if path.suffix.lower() == ".json":
        return services_from_json(path)
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return services_from_xlsx(path)
    raise SystemExit(f"Неподдерживаемый формат справочника: {path.suffix}")


def _match_all(norm: Normalizer, items: list):
    rows = []
    counts = {"auto": 0, "review": 0, "unmatched": 0}
    cache: dict[str, object] = {}        # одинаковые названия считаем один раз
    total = len(items)
    for n, it in enumerate(items, 1):
        raw = it.service_name_raw
        res = cache.get(raw)
        if res is None:
            res = norm.match(raw)
            cache[raw] = res
        if n % 1000 == 0 or n == total:
            print(f"  сопоставление… {n}/{total} (уникальных {len(cache)})", flush=True)
        if res.service_id and not res.needs_review:
            bucket = "auto"
        elif res.service_id and res.needs_review:
            bucket = "review"
        else:
            bucket = "unmatched"
        counts[bucket] += 1
        rows.append({
            "service_name_raw": raw,
            "matched_service": res.matched_name or "",
            "service_id": res.service_id or "",
            "confidence": res.confidence if res.confidence is not None else "",
            "stage": res.stage,
            "bucket": bucket,
            "price_resident_kzt": getattr(it, "price_resident_kzt", "") or "",
            "price_nonresident_kzt": getattr(it, "price_nonresident_kzt", "") or "",
        })
    return rows, counts


def _print_report(counts: dict, total: int, cfg: NormalizerConfig) -> None:
    auto_rate = counts["auto"] / total if total else 0.0
    print("=" * 56)
    print("Результат сопоставления (твои данные)")
    print("-" * 56)
    print(f"Пороги: auto={cfg.auto_threshold}  unmatched={cfg.unmatched_threshold}")
    print(f"Всего позиций:        {total}")
    print(f"  автосопоставлено:   {counts['auto']}")
    print(f"  на ручную проверку: {counts['review']}")
    print(f"  не сопоставлено:    {counts['unmatched']}")
    print(f"Доля автосопоставл.:  {auto_rate:.1%}  (цель ТЗ ≥ 70%)")
    print("=" * 56)


def _write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    ap = argparse.ArgumentParser(description="Локальный тест пайплайна на своих данных")
    ap.add_argument("--archive", help="ZIP, папка или один файл с прайсами")
    ap.add_argument("--inspect-services", action="store_true",
                    help="Только проверить справочник (без парсинга архива)")
    ap.add_argument("--services", help="Справочник: JSON или XLSX")
    ap.add_argument("--from-db", action="store_true", help="Взять справочник из Supabase")
    ap.add_argument("--out", default="pipeline_test.csv", help="CSV с детализацией")
    ap.add_argument("--auto", type=float, default=None)
    ap.add_argument("--unmatched", type=float, default=None)
    ap.add_argument("--no-embeddings", action="store_true",
                    help="Быстрый прогон без LaBSE (только точное/синонимы/fuzzy)")
    run(ap.parse_args())


if __name__ == "__main__":
    main()
